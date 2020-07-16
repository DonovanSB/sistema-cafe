#!/usr/bin/env python
# -*- coding: utf-8 -*-
import random
from openpyxl import Workbook, load_workbook
from datetime import datetime
import matplotlib.dates as dates
import os
from PyQt5.QtCore import pyqtSignal, QObject
import json
import logging
import paho.mqtt.client as mqtt 

rutaPrefsUser = os.path.dirname(os.path.abspath(__file__))

parent = os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir, os.pardir)
root = os.path.abspath(parent)
routeDatos = root + '/datos'

logging.basicConfig(filename = root + '/fermentacion.log', format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

def singleton(cls):    
    instance = [None]
    def wrapper(*args, **kwargs):
        if instance[0] is None:
            instance[0] = cls(*args, **kwargs)
        return instance[0]

    return wrapper

class Data:
    def __init__(self,textTempR,textHumR,textTemp1,textTemp2,textTemp3,textBrix,textPh):
        
        self.textTempR = textTempR
        self.textHumR  = textHumR
        self.textTemp1 = textTemp1
        self.textTemp2 = textTemp2
        self.textTemp3 = textTemp3
        self.textBrix  = textBrix
        self.textPh    = textPh

        self.numData   = 1000

        self.prefs = LocalStorage(route=rutaPrefsUser, name = 'prefs')
        try:
            self.routeData = os.path.abspath(self.prefs.read()['routeData'] + '/datos.xlsx')
        except:
            logging.error('No se pudo encontrar la ruta de almacenamiento de datos en prefs.json')
            self.routeData = routeDatos + '/datos.xlsx'
       
        # Inicializaciones para almacenamiento de datos
        self.wb = WBook(self.routeData).workbook
        self.initExcel(self.routeData)
        self.initDataService()

        self.signals = Signals()
        self.signals.signalUpdatePrefs.connect(self.updatePrefs)
        self.signals.signalUpdateInputValue.connect(self.updateInputValue)

        # Mqqt
        self.client = Mqtt('fermentacion')

    def initDataService(self):
        self.envService = DataService( ['T.ambiente', 'H.ambiente'], [self.textTempR, self.textHumR], self.envExcel, self.numData, ["°C", "%"], 2)
        self.temp1Service = DataService( 'Temp1', self.textTemp1, self.temp1Excel, self.numData, "°C")
        self.temp2Service = DataService( 'Temp2', self.textTemp2, self.temp2Excel, self.numData, "°C")
        self.temp3Service = DataService( 'Temp3', self.textTemp3, self.temp3Excel, self.numData, "°C")
        self.brixService = DataService( 'Brix', self.textBrix, self.brixExcel, self.numData, "Bx")
        self.phService = DataService( 'ph', self.textPh, self.temp3Excel, self.numData, "")

    def initExcel(self, route):
        self.envExcel = Excel(wb = self.wb, titleSheet='Ambiente', head = ['Tiempo', 'Temperatura', 'Humedad'], route = route, initial = True)
        self.temp1Excel = Excel(wb = self.wb, titleSheet='Temperatura 1', head = ['Tiempo', 'Temperatura'], route = route)
        self.temp2Excel = Excel(wb = self.wb, titleSheet='Temperatura 2', head = ['Tiempo', 'Temperatura'], route = route)
        self.temp3Excel = Excel(wb = self.wb, titleSheet='Temperatura 3', head = ['Tiempo', 'Temperatura'], route = route)
        self.brixExcel = Excel(wb = self.wb, titleSheet='Brix', head = ['Tiempo', 'Brix'], route = route)
        self.phExcel = Excel(wb = self.wb, titleSheet='PH', head = ['Tiempo', 'PH'], route = route)

    def updateInputValue(self,name,value):
        currentTime = datetime.now()
        if name == "Brix":
            self.brixService.update(value, currentTime)
        if name == "PH":
            self.phService.update(value, currentTime)

    def updatePrefs(self,route):
        self.routeData = os.path.abspath(route+ '/datos.xlsx')
        self.wb = WBook(self.routeData).workbook
        self.initExcel(self.routeData)
        self.initDataService()
        self.client.connect()

    def getData(self, index):
        datos= [self.envService.data[0], self.envService.data[1], self.temp1Service.data, self.temp2Service.data, self.temp3Service.data, self.brixService.data, self.phService.data]     
        return datos[index]
    
    def getTime(self, index):
        time = [self.envService.time, self.envService.time, self.temp1Service.time, self.temp2Service.time, self.temp3Service.time, self.brixService.time, self.phService.time]
        return time[index]

    #  ************** Lectura, visualización y almacenamiento ************
    def env(self):
        # Leer datos
        temperatureR = random.randint(18, 25)
        humidityR = random.randint(50, 60)
        currentTime = datetime.now()
        self.envService.update([temperatureR, humidityR], currentTime)

    def temp1(self):
        # Leer Datos
        temperature = random.randint(18, 25)
        currentTime = datetime.now()
        self.temp1Service.update(temperature, currentTime)

    def temp2(self):
        # Leer Datos
        temperature = random.randint(18, 25)
        currentTime = datetime.now()
        self.temp2Service.update(temperature, currentTime)

    def temp3(self):
        # Leer Datos
        temperature = random.randint(18, 25)
        currentTime = datetime.now()
        self.temp3Service.update(temperature, currentTime)
    
    def save(self):
        try:
            self.wb.save(self.routeData)
        except:
            logging.error('Ingrese un ruta correcta para almacenar los datos')
            print("provider save: Ingrese un ruta correcta para almacenar los datos")
        
        self.signals.statusFile = True
            
class DataService:
    def __init__(self, name, text, excel, numData, units,numVarSensor = 1):
        self.name = name
        self.text = text
        self.excel = excel
        self.numData = numData
        self.numVarSensor = numVarSensor
        self.units = units
        self.signals = Signals()

        # Mqqt
        self.client = Mqtt('fermentacion')

        if self.numVarSensor > 1:
            self.data = [[] for i in range(self.numVarSensor)]
        else:
            self.data = []
        self.time = []
    
    def update(self, data, time):
        self.time.append(time)
        if len(self.time) > self.numData:
            self.time.pop(0)

        if self.numVarSensor > 1:
            for i in range(self.numVarSensor):
                self.data[i].append(data[i])
                if len(self.data[i]) > self.numData:
                    self.data[i].pop(0)
                self.text[i].setText( str(data[i]) +" "+ self.units[i])
                # Enviar datos al servidor
                self.client.publish(json.dumps({self.name[i]:data[i], 'time':str(time)}))
            datos = [time]
            datos.extend(data)
            self.excel.addRow(datos)
        else:
            self.data.append(data)
            if len(self.data) > self.numData:
                self.data.pop(0)
            self.text.setText( str(data) +" "+ self.units)
            self.excel.addRow([time, data])
            # Enviar datos al servidor
            self.client.publish(json.dumps({self.name:data, 'time':str(time)}))

        # Actualizar grafica  
        self.signals.signalUpdateGraph.emit()
        
@singleton
class Mqtt:
    def __init__(self, clientID):
        self.prefs = LocalStorage(route=rutaPrefsUser, name = 'prefs')
        self.pendingData = LocalStorage(route=routeDatos, name = 'pending')
        self.client = mqtt.Client(client_id=clientID, clean_session=True, userdata=None, transport="tcp")
        self.isPending = False
        self.data = []
        try:
            self.topic = self.prefs.read()['topic']
        except:
            logging.error('No se encontró topic en prefs.json')
            self.topic = 'estacion/secado'
        self.connect()

    def connect(self):
        try:
            self.brokerAddress = self.prefs.read()['server']
        except:
            logging.error('No se encontró server en prefs.json')
        try:
            self.client.username_pw_set(username="usuario_publicador_1",password="123")
            if self.client.is_connected:
                self.client.disconnect()
                self.client.connect(self.brokerAddress, port=1884)
            else:
                self.client.connect(self.brokerAddress, port=1884)
        except:
            logging.error('No se pudo conectar al servidor')
    
    def publish(self, payload):
        info = self.client.publish(self.topic, payload)
        if info.is_published() == False:
            logging.error('No se pudo publicar los datos en el servidor')
            if self.pendingData.read():
                self.data = self.pendingData.read()
            self.data.append(payload)
            self.pendingData.update(self.data)
            self.isPending = True
    
    def verifyPending(self):
        if self.isPending:
            data = self.pendingData.read()
            dataTemp = data
            if data:
                for item in dataTemp:
                    try:
                        self.client.publish(self.topic, item)
                        data.remove(item)
                    except:
                        logging.error('No se pudo publicar los datos en el servidor')
                self.pendingData.update(data)
            else:
                self.isPending = False      

class Plotter:
    def __init__(self,Figure,ax):
        self.Fig = Figure
        self.ax = ax
        self.formatTime = dates.DateFormatter("%H:%M")
        self.limits = [[0,80], [0,100], [0,80], [0,80], [0,80], [0,85], [0,14]]

    def plot(self,datax,datay,title,index):
        if len(datax) <= 0 and len(datay) <= 0:
            datax = [datetime.now()]
            datay = [0]
        self.ax.cla()
        if len(datax) == len(datay):
            self.ax.plot(datax,datay, 'b-')
        self.ax.xaxis.set_major_formatter(self.formatTime)
        self.ax.grid()
        self.ax.set_ylim(self.limits[index][0],self.limits[index][1])
        self.ax.set_title(title,fontsize = "18", fontweight='bold')
        self.ax.set_xlabel('Hora',fontsize = "15")
        self.ax.set_ylabel('x(t)',fontsize = "15")
        self.Fig.figure.subplots_adjust(top = 0.85,bottom=0.2, left=0.1, right = 0.95)
        self.Fig.draw()

class WBook:
    def __init__(self, route):

        signals = Signals()

        fileExists = os.path.isfile(route)
        if fileExists:
            try:
                self.workbook = load_workbook(filename= route)
            except:
                logging.error('Archivo de respaldo de datos dañado')
                print('Archivo de respaldo de datos dañado')
                signals.statusFile = False
                self.workbook = Workbook()
        else:
            self.workbook = Workbook()

class Excel:
    def __init__(self, wb, titleSheet, head, route, initial = False):

        signals = Signals()

        fileExists = os.path.isfile(route)
        self.workbook = wb

        if fileExists and signals.statusFile:
            self.sheet = self.workbook.get_sheet_by_name(titleSheet)
        else:
            if initial:
                self.sheet = self.workbook.active
                self.sheet.title = titleSheet
            else:
                self.sheet = self.workbook.create_sheet(title=titleSheet)  
            self.sheet.append(head)

    def addRow(self, data):
        self.sheet.append(data)
     

class LocalStorage():
    def __init__(self, route, name):
        self.optionsServer = []
        self.routePrefs = os.path.abspath(route + "/" + name + ".json")
        self.name = name

    def read(self):
        try:
            with open(self.routePrefs) as file:
                if file.read():
                    file = open(self.routePrefs)
                    return json.load(file)
                else:
                    logging.error('No se encontraron las preferencias del usuario')
                    print('No se encontraron las preferencias del usuario')
                    return False
        except:
            logging.error('No se pudo encontrar el archivo ' + self.name + '.json')
            print('No se pudo encontrar el archivo ' + self.name + '.json')
            return False

    def update(self,prefs):
        with open(self.routePrefs, 'w') as file:
            json.dump(prefs, file)


@singleton
class Signals(QObject):
    signalUpdatePrefs = pyqtSignal(str)
    signalUpdateInputValue = pyqtSignal(float)
    signalUpdateGraph = pyqtSignal()

    statusFile = True
    dataPending = False
