#!/usr/bin/env python
# -*- coding: utf-8 -*-
import random
from openpyxl import Workbook, load_workbook
from datetime import datetime
import matplotlib.dates as dates
import os
from PyQt5.QtCore import pyqtSignal, QObject
import json

route = os.path.dirname(os.getcwd()) + "/Estacion Cafe Secado"
rutaPrefsUser = route + "/src/providers"

class Data:
    def __init__(self,textHumGrain, textTempEnv, textHumEnv,textTemp1,textHum1,textTemp2,textHum2,textTemp3,textHum3):
        
        self.textTempEnv  = textTempEnv
        self.textHumEnv   = textHumEnv
        self.textHumGrain = textHumGrain
        self.textTemp1    = textTemp1
        self.textHum1     = textHum1
        self.textTemp2    = textTemp2
        self.textHum2     = textHum2
        self.textTemp3    = textTemp3
        self.textHum3     = textHum3
        
        
        self.humGrainData = []
        self.tempEnvData  = []
        self.humEnvData   = []
        self.temp1Data    = []
        self.hum1Data     = []
        self.temp2Data    = []
        self.hum2Data     = []
        self.temp3Data    = []
        self.hum3Data     = []

        self.envTime      = []
        self.env1Time     = []
        self.env2Time     = []
        self.env3Time     = []
        self.humGrainTime = []
       
        self.numData      = 1000

        self.humidityGrain = 50

        self.prefs = LocalStorage()
        self.prefs.beginPrefs(route=rutaPrefsUser)
        prefs = self.prefs.readPrefs()

        # Inicializaciones para almacenamiento de datos
        self.wb = WBook(prefs['routeData']).workbook
        self.initExcel(prefs['routeData'])

        self.signals = Signals()
        self.signals.signalUpdateStorageRoute.connect(self.updateStorgeRoute)
        self.signals.signalUpdateInputValue.connect(self.updateInputValue)
    
    def initExcel(self, route):
        self.humGrainExcel = Excel(wb = self.wb, titleSheet='Humedad Grano', head = ['Tiempo', 'Humedad Grano'], route = route, initial = True)
        self.envExcel = Excel(wb = self.wb, titleSheet='Ambiente', head = ['Tiempo', 'Temperatura', 'Humedad'], route = route)
        self.env1Excel = Excel(wb = self.wb, titleSheet='Zona 1', head = ['Tiempo', 'Temperatura', 'Humedad'], route = route)
        self.env2Excel = Excel(wb = self.wb, titleSheet='Zona 2', head = ['Tiempo', 'Temperatura', 'Humedad'], route = route)
        self.env3Excel = Excel(wb = self.wb, titleSheet='Zona 3', head = ['Tiempo', 'Temperatura', 'Humedad'], route = route)

    def updateInputValue(self,hum):
        self.humidityGrain = hum
        currentTime = datetime.now()
        self.humGrainData.append(self.humidityGrain)
        self.humGrainTime.append(currentTime)
        if len(self.humGrainData) > self.numData:
            self.humGrainData.pop(0)
        if len(self.humGrainTime) > self.numData:
            self.humGrainTime.pop(0)

        self.signals.signalUpdateGraph.emit()
        self.textHumGrain.setText(str(self.humidityGrain) + " %")

        self.humGrainExcel.save([currentTime, self.humidityGrain])

    def updateStorgeRoute(self,route):
        self.wb = WBook(route).workbook
        self.initExcel(route)

    def getData(self, index):
        self.datos= [self.humGrainData,self.tempEnvData,self.humEnvData,self.temp1Data,self.hum1Data,self.temp2Data,self.hum2Data,self.temp3Data,self.hum3Data]     
        return self.datos[index]

    def getTime(self, index):
        self.time = [self.humGrainTime,self.envTime,self.envTime,self.env1Time,self.env1Time,self.env2Time,self.env2Time,self.env3Time,self.env3Time]
        return self.time[index]

    def env(self):
        self.temperatureEnv = random.randint(18, 25)
        self.humidityEnv = random.randint(50, 60)
        currentTime = datetime.now()

        self.tempEnvData.append(self.temperatureEnv)
        self.humEnvData.append(self.humidityEnv)
        self.envTime.append(currentTime)

        if len(self.tempEnvData) > self.numData:
            self.tempEnvData.pop(0)
        
        if len(self.humEnvData) > self.numData:
            self.humEnvData.pop(0)

        if len(self.envTime) > self.numData:
            self.envTime.pop(0)

        self.signals.signalUpdateGraph.emit()
        self.textTempEnv.setText( str(self.temperatureEnv) + " °C")
        self.textHumEnv.setText( str(self.humidityEnv) + " %")

        self.envExcel.save([currentTime, self.temperatureEnv, self.humidityEnv])
    
    def env1(self):
        self.temperature1 = random.randint(18, 25)
        self.humidity1 = random.randint(50, 60)
        currentTime = datetime.now()
        
        self.temp1Data.append(self.temperature1)    
        self.hum1Data.append(self.humidity1)
        self.env1Time.append(currentTime)

        if len(self.temp1Data) > self.numData:
            self.temp1Data.pop(0)
        
        if len(self.hum1Data) > self.numData:
            self.hum1Data.pop(0)
        
        if len(self.env1Time) > self.numData:
            self.env1Time.pop(0)

        self.signals.signalUpdateGraph.emit()
        self.textTemp1.setText( str(self.temperature1) + " °C")
        self.textHum1.setText( str(self.humidity1) + " %")

        self.env1Excel.save([currentTime, self.temperature1, self.humidity1])
    
    def env2(self):
        self.temperature2 = random.randint(18, 25)
        self.humidity2 = random.randint(50, 60)
        currentTime = datetime.now()
        
        self.temp2Data.append(self.temperature2)    
        self.hum2Data.append(self.humidity2)
        self.env2Time.append(currentTime)
        
        if len(self.temp2Data) > self.numData:
            self.temp2Data.pop(0)
        
        if len(self.hum2Data) > self.numData:
            self.hum2Data.pop(0)
        
        if len(self.env2Time) > self.numData:
            self.env2Time.pop(0)

        self.signals.signalUpdateGraph.emit()
        self.textTemp2.setText( str(self.temperature2) + " °C")
        self.textHum2.setText( str(self.humidity2) + " %")

        self.env2Excel.save([currentTime, self.temperature2, self.humidity2])

    def env3(self):
        self.temperature3 = random.randint(18, 25)
        self.humidity3 = random.randint(50, 60)
        currentTime = datetime.now()
        
        self.temp3Data.append(self.temperature3)    
        self.hum3Data.append(self.humidity3)
        self.env3Time.append(currentTime)
        
        if len(self.temp3Data) > self.numData:
            self.temp3Data.pop(0)
        
        if len(self.hum3Data) > self.numData:
            self.hum3Data.pop(0)
        
        if len(self.env3Time) > self.numData:
            self.env3Time.pop(0)

        self.signals.signalUpdateGraph.emit()
        self.textTemp3.setText( str(self.temperature3) + " °C")
        self.textHum3.setText( str(self.humidity3) + " %")

        self.env3Excel.save([currentTime, self.temperature3, self.humidity3])


class Plotter:
    def __init__(self,Figure,ax):
        self.Fig = Figure
        self.ax = ax
        self.formatTime = dates.DateFormatter("%H:%M")
        self.limits = [[0,100], [0,80], [0,100], [0,80], [0,100], [0,80], [0,100], [0,80], [0,100]]

    def plot(self,datax,datay,title,index):
        if len(datax) <= 0 and len(datay) <= 0:
            datax = datetime.now()
            datay = 0
        self.ax.cla()
        self.ax.plot(datax,datay, 'b-')
        self.ax.xaxis.set_major_formatter(self.formatTime)
        self.ax.grid()
        self.ax.set_ylim(self.limits[index][0],self.limits[index][1])
        self.ax.set_title(title,fontsize = "18", fontweight='bold')
        self.ax.set_xlabel('Hora',fontsize = "15")
        self.ax.set_ylabel('x(t)',fontsize = "15")
        self.Fig.figure.subplots_adjust(top = 0.8,bottom=0.27, left=0.15)
        self.Fig.draw()

class WBook:
    def __init__(self, route):
        fileExists = os.path.isfile(route + '/datos.xlsx')
        if fileExists:
            self.workbook = load_workbook(filename= route + '/datos.xlsx')
        else:
            self.workbook = Workbook()

class Excel:
    def __init__(self, wb, titleSheet, head, route, initial = False):
        fileExists = os.path.isfile(route + '/datos.xlsx')
        self.route = route + '/'
        self.workbook = wb

        if fileExists:
            self.sheet = self.workbook.get_sheet_by_name(titleSheet)
        else:
            if initial:
                self.sheet = self.workbook.active
                self.sheet.title = titleSheet
            else:
                self.sheet = self.workbook.create_sheet(title=titleSheet)  
            self.sheet.append(head)
        # print(len(tuple(self.sheet.rows)))
        # print(self.sheet.cell(row=3, column=1).value)

    def save(self, data):
        self.sheet.append(data)
        try:
            self.workbook.save(self.route+ 'datos' + '.xlsx')
        except:
            print("provider save: Ingrese un ruta correcta para almacenar los datos")
            
class LocalStorage():
    def __init__(self):
        self.optionsServer = []
    
    def beginPrefs(self,route):
        self.routePrefs = route + "/"
    
    def readPrefs(self):
        try:
            with open(self.routePrefs + 'prefs.json') as file:
                if file.read():
                    file = open(self.routePrefs + 'prefs.json')
                    return json.load(file)
                else:
                    print('No se encontraron las preferencias del usuario')
                    return False
        except:
            print('No se pudo encontrar el archivo de prefs.json')
            return False

    def updatePrefs(self,prefs):
        with open(self.routePrefs + 'prefs.json', 'w') as file:
            json.dump(prefs, file)

def singleton(cls):    
    instance = [None]
    def wrapper(*args, **kwargs):
        if instance[0] is None:
            instance[0] = cls(*args, **kwargs)
        return instance[0]

    return wrapper
@singleton
class Signals(QObject):
    signalServerUpdate = pyqtSignal(bool)
    signalUpdateStorageRoute = pyqtSignal(str)
    signalUpdateInputValue = pyqtSignal(float)
    signalUpdateGraph = pyqtSignal()
