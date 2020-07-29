#!/usr/bin/env python
# -*- coding: utf-8 -*-
import random
from datetime import datetime
import matplotlib.dates as dates
from openpyxl import Workbook, load_workbook
import os
from PyQt5.QtCore import pyqtSignal, QObject, QThread
import json
import logging
import paho.mqtt.client as mqtt 

rutaPrefsUser = os.path.dirname(os.path.abspath(__file__))

parent = os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir, os.pardir)
root = os.path.abspath(parent)
routeDatos = root + '/datos'
logging.basicConfig(filename = root + '/estacion.log', format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

def singleton(cls):    
    instance = [None]
    def wrapper(*args, **kwargs):
        if instance[0] is None:
            instance[0] = cls(*args, **kwargs)
        return instance[0]

    return wrapper

class Data:
    def __init__(self,loading,textTemp,textHum,textIrrad,textSpeed,textDir,textRain):
        self.loading = loading
        self.textTemp = textTemp
        self.textHum = textHum
        self.textIrrad = textIrrad
        self.textSpeed = textSpeed
        self.textDir = textDir
        self.textRain = textRain

        self.numData = 1000

        self.prefs = LocalStorage(route=rutaPrefsUser, name = 'prefs')
        try:
            self.routeData = os.path.abspath(self.prefs.read()['routeData'] + '/datos.xlsx')
        except:
            logging.error('No se pudo encontrar la ruta de almacenamiento de datos en prefs.json')
            self.routeData = routeDatos + '/datos.xlsx'
        
        # Inicializaciones para almacenamiento de datos
        self.wb = WBook(self.routeData).workbook
        self.initExcel(self.routeData)
        self.initDataService()

        self.signals = Signals()
        self.signals.signalIsLoanding.connect(self.isLoading)
        self.signals.signalUpdatePrefs.connect(self.updatePrefs)

        # Mqqt
        self.client = Mqtt('estacion')
    
    def initDataService(self):
        self.envService = DataService(['T.ambiente', 'H.ambiente'], [self.textTemp, self.textHum], self.envExcel, self.numData, ["°C", "%"], 2)
        self.irradService = DataService( 'Irrad', self.textIrrad, self.irradExcel, self.numData, "W/m²")
        self.speedService = DataService( 'Velocidad', self.textSpeed, self.speedExcel, self.numData, "km/h")
        self.directionService = DataService( 'Direccion', self.textDir, self.directionExcel, self.numData, "°")
        self.rainService = DataService( 'Lluvia', self.textRain, self.rainExcel, self.numData, "cm/h")

    def initExcel(self, route):
        self.envExcel = Excel(wb = self.wb, titleSheet='Ambiente', head = ['Tiempo', 'Temperatura', 'Humedad'], route = route, initial = True)
        self.irradExcel = Excel(wb = self.wb, titleSheet='Irradiancia', head = ['Tiempo', 'Irradiancia'], route = route)
        self.speedExcel = Excel(wb = self.wb, titleSheet='Velocidad V', head = ['Tiempo', 'Velocidad'], route = route)
        self.directionExcel = Excel(wb = self.wb, titleSheet='Direccion V', head = ['Tiempo', 'Direccion'], route = route)
        self.rainExcel = Excel(wb = self.wb, titleSheet='Lluvia', head = ['Tiempo', 'Lluvia'], route = route)

    def isLoading(self, loading):
        if loading:
            self.loading.start()
        else:
            self.loading.stop()
    
    def updatePrefs(self,route):
        self.routeData = os.path.abspath(route + '/datos.xlsx')
        self.wb = WBook(self.routeData).workbook
        self.initExcel(self.routeData)
        self.initDataService()
        self.thread = Thread(target = self.client.connect)
        self.thread.start()
    
    def getData(self, index):
        data = [self.envService.data[0], self.envService.data[1], self.irradService.data, self.speedService.data, self.directionService.data, self.rainService.data]
        return data[index]

    def getTime(self, index):
        time = [self.envService.time, self.envService.time, self.irradService.time, self.speedService.time, self.directionService.time, self.rainService.time]
        return time[index]

    def env(self):
        # Leer datos
        temperatureR = random.randint(18, 25)
        humidityR = random.randint(50, 60)
        currentTime = datetime.now()
        self.envService.update([temperatureR, humidityR], currentTime)
    
    def irrad(self):
        irradiance = random.randint(200, 300)
        currentTime = datetime.now()
        self.irradService.update(irradiance, currentTime)

    def windSpeed(self):
        speed = random.randint(5, 10)
        currentTime = datetime.now()
        self.speedService.update(speed, currentTime)

    def windDirection(self):
        direction = random.randint(0, 360)
        currentTime = datetime.now()
        self.directionService.update(direction, currentTime)

    def rain(self):
        rain = random.randint(0, 5)
        currentTime = datetime.now()
        self.rainService.update(rain, currentTime)

    def save(self):
        try:
            self.wb.save(self.routeData)
        except:
            logging.error('Ingrese un ruta correcta para almacenar los datos')
            print("provider save: Ingrese un ruta correcta para almacenar los datos")
        
        self.signals.statusFile = True

class DataService:
    def __init__(self, name, text, excel, numData, units,numVarSensor = 1):
        self.name = name
        self.text = text
        self.excel = excel
        self.numData = numData
        self.numVarSensor = numVarSensor
        self.units = units
        self.signals = Signals()

        # Mqqt
        self.client = Mqtt('estacion')

        if self.numVarSensor > 1:
            self.data = [[] for i in range(self.numVarSensor)]
        else:
            self.data = []
        self.time = []
    
    def update(self, data, time):
        self.time.append(time)
        if len(self.time) > self.numData:
            self.time.pop(0)

        if self.numVarSensor > 1:
            for i in range(self.numVarSensor):
                self.data[i].append(data[i])
                if len(self.data[i]) > self.numData:
                    self.data[i].pop(0)
                self.text[i].setText( str(data[i]) +" "+ self.units[i])
                # Enviar datos al servidor
                self.client.publish(json.dumps({self.name[i]:data[i], 'time':str(time)}))
            datos = [time]
            datos.extend(data)
            self.excel.addRow(datos)
        else:
            self.data.append(data)
            if len(self.data) > self.numData:
                self.data.pop(0)
            self.text.setText( str(data) +" "+ self.units)
            self.excel.addRow([time, data])
            # Enviar datos al servidor
            self.client.publish(json.dumps({self.name:data, 'time':str(time)}))

        # Actualizar grafica  
        self.signals.signalUpdateGraph.emit()

@singleton
class Mqtt:
    def __init__(self, clientID):
        self.prefs = LocalStorage(route=rutaPrefsUser, name = 'prefs')
        self.pendingData = LocalStorage(route=routeDatos, name = 'pending')
        self.client = mqtt.Client(client_id=clientID, clean_session=True, userdata=None, transport="tcp")
        self.isPending = True
        self.isConnected = False
        if self.pendingData.read():
            self.data = self.pendingData.read()
        else:
            self.data = []

        self.signals = Signals()
        try:
            self.topic = self.prefs.read()['topic']
        except:
            logging.error('No se encontró topic en prefs.json')
            self.topic = 'estacion/estacion'

        # Conectar en segundo plano
        self.thread = Thread(target = self.connect)
        self.thread.start()

    def connect(self):
        self.signals.signalIsLoanding.emit(True)
        try:
            self.brokerAddress = self.prefs.read()['server']
        except:
            logging.error('No se encontró server en prefs.json')
        try:
            self.client.username_pw_set(username="usuario_publicador_1",password="123")
            if self.client.is_connected:
                self.client.disconnect()
                self.client.connect(self.brokerAddress, port=1884)
            else:
                self.client.connect(self.brokerAddress, port=1884)
            
            self.isConnected = True
        except:
            self.signals.signalAlert.emit('No se pudo conectar al servidor')
            logging.error('No se pudo conectar al servidor')
            self.isConnected = False
        
        self.signals.signalIsLoanding.emit(False)
    
    def publish(self, payload):
        info = self.client.publish(self.topic, payload)
        if info.is_published() == False:
            logging.error('No se pudo publicar los datos en el servidor')
            if self.pendingData.read():
                self.data = self.pendingData.read()
            self.data.append(payload)
            self.pendingData.update(self.data)
            self.isPending = True
            self.isConnected = False
        else:
            self.isConnected = True
    
    def verifyPending(self):
        if self.isPending and self.isConnected:
            dataTemp = self.data
            if dataTemp:
                for item in dataTemp:
                    info = self.client.publish(self.topic, item)
                    if info.is_published():
                        self.data.remove(item)
                        self.pendingData.update(self.data)
                    else:
                        logging.error('No se pudo publicar los datos en el servidor')  
                        self.isConnected = False   
            else:
                self.isPending = False                            

class Plotter:
    def __init__(self,Figure,ax):
        self.Fig = Figure
        self.ax = ax
        self.formatTime = dates.DateFormatter("%H:%M")
        self.limits = [[0,80], [0,100], [0,1200], [0,100], [0,360], [0,50]]

    def plot(self,datax,datay,title,index):
        if len(datax) <= 0 and len(datay) <= 0:
            datax = [datetime.now()]
            datay = [0]
        self.ax.cla()
        if len(datax) == len(datay):
            self.ax.plot(datax,datay, 'b-')
        self.ax.plot(datax,datay, 'b-')
        self.ax.xaxis.set_major_formatter(self.formatTime)
        self.ax.grid()
        self.ax.set_ylim(self.limits[index][0],self.limits[index][1])
        self.ax.set_title(title,fontsize = "18", fontweight='bold')
        self.ax.set_xlabel('Hora',fontsize = "15")
        self.ax.set_ylabel('x(t)',fontsize = "15")
        self.Fig.figure.subplots_adjust(top = 0.85,bottom=0.2, left=0.13, right = 0.95)
        self.Fig.draw()

class WBook:
    def __init__(self, route):

        signals = Signals()

        fileExists = os.path.isfile(route)
        if fileExists:
            try:
                self.workbook = load_workbook(filename= route)
            except:
                logging.error('Archivo de respaldo de datos dañado')
                print('Archivo de respaldo de datos dañado')
                signals.statusFile = False
                self.workbook = Workbook()
        else:
            self.workbook = Workbook()

class Excel:
    def __init__(self, wb, titleSheet, head, route, initial = False):

        signals = Signals()

        fileExists = os.path.isfile(route)
        self.workbook = wb

        if fileExists and signals.statusFile:
            self.sheet = self.workbook.get_sheet_by_name(titleSheet)
        else:
            if initial:
                self.sheet = self.workbook.active
                self.sheet.title = titleSheet
            else:
                self.sheet = self.workbook.create_sheet(title=titleSheet)  
            self.sheet.append(head)

    def addRow(self, data):
        self.sheet.append(data)

class LocalStorage():
    def __init__(self, route, name):
        self.optionsServer = []
        self.routePrefs = os.path.abspath(route + "/" + name + ".json")
        self.name = name

    def read(self):
        try:
            with open(self.routePrefs) as file:
                if file.read():
                    file = open(self.routePrefs)
                    return json.load(file)
                else:
                    logging.error('No se encontraron las preferencias del usuario')
                    print('No se encontraron las preferencias del usuario')
                    return False
        except:
            logging.error('No se pudo encontrar el archivo ' + self.name + '.json')
            print('No se pudo encontrar el archivo ' + self.name + '.json')
            return False

    def update(self,prefs):
        with open(self.routePrefs, 'w') as file:
            json.dump(prefs, file)


class Thread(QThread):

    def __init__(self, target):
        super(Thread,self).__init__()
        self.threadactive = True
        self.target = target

    def run(self):
        self.target()
    
    def stop(self):
        self.threadactive = False

@singleton
class Signals(QObject):
    signalUpdatePrefs = pyqtSignal(str)
    signalUpdateInputValue = pyqtSignal(float)
    signalUpdateGraph = pyqtSignal()
    signalIsLoanding = pyqtSignal(bool)
    signalAlert = pyqtSignal(str)

    statusFile = True
    dataPending = False
